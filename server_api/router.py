from flask import Flask, send_file, render_template, request, jsonify
import oracledb
import pandas as pd
from datetime import datetime, date
import io
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
from flask_cors import CORS
from waitress import serve
from dotenv import load_dotenv
import os
import logging
from dotenv import load_dotenv

logging.basicConfig(level=logging.DEBUG, format='%(asctime)s %(message)s')


load_dotenv('.env')

app = Flask(__name__)  # Configura a pasta estática
CORS(app, resources={r"/report": {"origins": "*"}})

# @app.route('/')
# def index():
#     return render_template('index.html')  # Renderiza o index.html

@app.route("/report", methods=['GET'])
def report():
    ref_ext = request.args.get('ref_ext')
    tipo_nota = request.args.get('tipo_nota')
    data_de = request.args.get('data_de')
    data_ate = request.args.get('data_ate')
    
    #VALIDAÇÕES
    if tipo_nota is None:
        return jsonify({"error": "Parâmetro 'tipo_nota' é obrigatório"}), 400

    # Tratamento do tipo_nota
    tipo_nota_list = []
    try:
        for tn in tipo_nota.split(','):
            if tn.strip():
                tipo_nota_list.append(int(tn.strip()))
        if not tipo_nota_list:
            return jsonify({"error": "Pelo menos um tipo de nota válido deve ser fornecido"}), 400
        tipo_nota_str = ','.join(map(str, tipo_nota_list))
    except ValueError:
        return jsonify({"error": "Formato inválido para 'tipo_nota'. Use números inteiros separados por vírgula."}), 400

    print("Tipos de nota processados:", tipo_nota_list)
    print("String de tipos de nota para SQL:", tipo_nota_str)

    if not ref_ext or ref_ext.lower() == 'null':
        ref_ext_condition = "IS NOT NULL"
        ref_ext_param = {}
    else:
        ref_ext_list = ref_ext.split(',')
        ref_ext_condition = "IN ({})".format(','.join([':ref_ext{}'.format(i) for i in range(len(ref_ext_list))]))
        ref_ext_param = {f'ref_ext{i}': val.strip() for i, val in enumerate(ref_ext_list)}


    if data_de is None:
        hoje = date.today()
        data_de = date(hoje.year, hoje.month, 1).strftime('%d/%m/%Y')
    if data_ate is None:
        data_ate = datetime.now().strftime('%d/%m/%Y')

    print("Parâmetros recebidos:", ref_ext, tipo_nota, data_de, data_ate)

    if not all([tipo_nota]):
        return jsonify({"error": "Parâmetros incompletos"}), 400

    try:
        connection = oracledb.connect(
            user=os.getenv('USER'),
            password=os.getenv('PASSWORD'),
            host=os.getenv('HOST'),
            port=os.getenv('PORT'),
            service_name = os.getenv('SERVICE_NAME')
        )
        
        cursor = connection.cursor()

        # Preparar os tipos de nota
        tipo_nota_list = tipo_nota.split(',') if tipo_nota else ['0', '1', '3', '5', '9']
        tipo_nota_str = ','.join(tipo_nota_list)

        sql = """
        """.format(ref_ext_condition, tipo_nota_str)

        params = {
            'data_de': data_de,
            'data_ate': data_ate,
            **ref_ext_param
        }

        print("Parâmetros para SQL:", params)

        cursor.execute(sql, params)
        results = cursor.fetchall()

        print(f"Número de resultados: {len(results)}")

        if not results:
            print("Nenhum resultado encontrado")
            return jsonify({"message": "Nenhum resultado encontrado"}), 404

        df = pd.DataFrame(results, columns=[desc[0] for desc in cursor.description])
        
        output = io.BytesIO()
        wb = formatacao(df)
        wb.save(output)
        output.seek(0)

        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name='relatorio_.xlsx'
        )

    except oracledb.DatabaseError as e:
        print(f"Erro de banco de dados: {e}")
        return jsonify({"error": f"Erro ao acessar o banco de dados: {str(e)}"}), 500
    except Exception as e:
        print(f"Erro inesperado: {str(e)}")
        return jsonify({"error": f"Ocorreu um erro inesperado: {str(e)}"}), 500
    finally:
        if 'connection' in locals() and connection is not None:
            connection.close()

def formatacao(df):
    wb = Workbook()
    ws_sql = wb.active
    ws_sql.title = "Rateio de produtos"

    # Função para escrever DataFrame em um sheet
    for r in dataframe_to_rows(df, index=False, header=True):
        ws_sql.append(r)

    def format_sheet(ws):
        for row in ws.iter_rows():
            for cell in row:
                cell.alignment = Alignment(horizontal='center', vertical='center')

        blue_fill = PatternFill(start_color="CCE5FF", end_color="CCE5FF", fill_type="solid")
        white_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

        for idx, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row)):
            if idx % 2 == 0:
                for cell in row:
                    cell.fill = blue_fill
            else:
                for cell in row:
                    cell.fill = white_fill

    # Aplica a formatação
    format_sheet(ws_sql)
    return wb

if __name__ == '__main__':
    # Configurações para produção
    host = '0.0.0.0'
    port = 5000
    threads = 2

    
    serve(app, host=host, port=port, threads=threads)

